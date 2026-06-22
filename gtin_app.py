# -*- coding: utf-8 -*-
"""
GTIN 코드 발급 양식 생성기 (Streamlit)
- detailinfo(쿠팡 상품정보) 엑셀에서 상품/옵션을 읽어
  GTIN 발급 템플릿(업로드 템플릿 시트) 형식으로 변환한다.
"""
import io
import os
import re
import json
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(page_title="GTIN 코드 생성기", layout="wide")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# 0. 참조 데이터 로드 (상품분류 / 색상코드 / 신발 사이즈코드)
# ---------------------------------------------------------------------------
@st.cache_data
def load_ref():
    with open(os.path.join(BASE_DIR, "ref_data.json"), encoding="utf-8") as f:
        return json.load(f)

REF = load_ref()
COLOR_LIST = REF["color"]                       # [{code, name(영문)}]
SHOE_LIST = REF["shoe_size"]                    # [{code, size}]

# ---------------------------------------------------------------------------
# 신발 상품분류 (코드 내장 — JSON에 shoe_class가 없어도 항상 사용 가능)
# ---------------------------------------------------------------------------
SHOE_CLASS_BUILTIN = [
    ("11020101", "패션잡화 > 남성신발 > 모카신/털신/보트슈즈 > 모카신/털신"),
    ("11020102", "패션잡화 > 남성신발 > 모카신/털신/보트슈즈 > 보트슈즈"),
    ("11020201", "패션잡화 > 남성신발 > 부츠/샌들 > 부츠"),
    ("11020202", "패션잡화 > 남성신발 > 부츠/샌들 > 샌들"),
    ("11020301", "패션잡화 > 남성신발 > 스니커즈/슬립온/슬리퍼 > 스니커즈"),
    ("11020302", "패션잡화 > 남성신발 > 스니커즈/슬립온/슬리퍼 > 슬립온"),
    ("11020303", "패션잡화 > 남성신발 > 스니커즈/슬립온/슬리퍼 > 슬리퍼"),
    ("11020401", "패션잡화 > 남성신발 > 운동화 > 러닝화"),
    ("11020402", "패션잡화 > 남성신발 > 운동화 > 보드화"),
    ("11020403", "패션잡화 > 남성신발 > 운동화 > 아쿠아슈즈"),
    ("11020404", "패션잡화 > 남성신발 > 운동화 > 워킹화"),
    ("11020405", "패션잡화 > 남성신발 > 운동화 > 캔버스화"),
    ("11020406", "패션잡화 > 남성신발 > 운동화 > 하이탑"),
    ("11020501", "패션잡화 > 남성신발 > 워커/월트화/구두 > 워커"),
    ("11020502", "패션잡화 > 남성신발 > 워커/월트화/구두 > 웰트화"),
    ("11020503", "패션잡화 > 남성신발 > 워커/월트화/구두 > 정장구두"),
    ("11060101", "패션잡화 > 여성신발 > 기능화 > 다이어트화"),
    ("11060199", "패션잡화 > 여성신발 > 기능화 > 기타기능화"),
    ("11060201", "패션잡화 > 여성신발 > 단화 > 로퍼"),
    ("11060202", "패션잡화 > 여성신발 > 단화 > 모카신/털신"),
    ("11060203", "패션잡화 > 여성신발 > 단화 > 스니커즈"),
    ("11060204", "패션잡화 > 여성신발 > 단화 > 슬립온"),
    ("11060205", "패션잡화 > 여성신발 > 단화 > 옥스퍼드화"),
    ("11060206", "패션잡화 > 여성신발 > 단화 > 플랫"),
    ("11060301", "패션잡화 > 여성신발 > 부츠 > 롱부츠"),
    ("11060302", "패션잡화 > 여성신발 > 부츠 > 미들부츠"),
    ("11060303", "패션잡화 > 여성신발 > 부츠 > 앵클/숏부츠"),
    ("11060501", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 글래디에이터샌들"),
    ("11060502", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 뮬"),
    ("11060503", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 스트랩샌들"),
    ("11060504", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 슬링백샌들"),
    ("11060505", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 아쿠아샌들"),
    ("11060506", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 웨지힐샌들"),
    ("11060507", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 젤리샌들"),
    ("11060508", "패션잡화 > 여성신발 > 샌들/슬리퍼 > 슬리퍼"),
    ("11060701", "패션잡화 > 여성신발 > 운동화 > 러닝화"),
    ("11060702", "패션잡화 > 여성신발 > 운동화 > 보드화"),
    ("11060703", "패션잡화 > 여성신발 > 운동화 > 아쿠아슈즈"),
    ("11060704", "패션잡화 > 여성신발 > 운동화 > 워킹화"),
    ("11060705", "패션잡화 > 여성신발 > 운동화 > 캔버스화"),
    ("11060706", "패션잡화 > 여성신발 > 운동화 > 하이탑"),
    ("11060801", "패션잡화 > 여성신발 > 워커 > 워커힐"),
    ("11060802", "패션잡화 > 여성신발 > 워커 > 플랫워커"),
    ("11060901", "패션잡화 > 여성신발 > 힐/펌프스 > 가보시힐"),
    ("11060902", "패션잡화 > 여성신발 > 힐/펌프스 > 스트랩힐"),
    ("11060903", "패션잡화 > 여성신발 > 힐/펌프스 > 슬링백힐"),
    ("11060904", "패션잡화 > 여성신발 > 힐/펌프스 > 웨딩슈즈"),
    ("11060905", "패션잡화 > 여성신발 > 힐/펌프스 > 웨지힐"),
    ("11060906", "패션잡화 > 여성신발 > 힐/펌프스 > 토오픈힐"),
    ("11060907", "패션잡화 > 여성신발 > 힐/펌프스 > 통굽힐"),
    ("11060908", "패션잡화 > 여성신발 > 힐/펌프스 > 펌프스"),
    ("11060909", "패션잡화 > 여성신발 > 힐/펌프스 > 하이힐"),
    ("11120101", "패션잡화 > 신발용품 > 신발깔창 > 신발깔창"),
    ("11160101", "패션잡화 > 유니섹스신발 > 유니섹스신발류 > 실내화"),
    ("11160199", "패션잡화 > 유니섹스신발 > 유니섹스신발류 > 기타 유니섹스신발류"),
]
# JSON에 shoe_class가 있으면 그걸, 없으면 내장본을 사용
_json_shoe = REF.get("shoe_class")
if _json_shoe:
    SHOE_CLASS_LIST = _json_shoe
else:
    SHOE_CLASS_LIST = [{"code": c, "name": n} for c, n in SHOE_CLASS_BUILTIN]

# 분류: 신발 분류로 제한. 표시라벨 -> 코드
CLASS_LABEL2CODE = {f'{c["code"]}  |  {c["name"]}': c["code"] for c in SHOE_CLASS_LIST}
# 색상: 영문명 -> 코드
COLOR_NAME2CODE = {c["name"]: c["code"] for c in COLOR_LIST}
COLOR_CODE2NAME = {c["code"]: c["name"] for c in COLOR_LIST}
# 신발 사이즈: 사이즈문자열 -> 코드
SHOE_SIZE2CODE = {s["size"]: s["code"] for s in SHOE_LIST}
SHOE_CODE2SIZE = {s["code"]: s["size"] for s in SHOE_LIST}

# ---------------------------------------------------------------------------
# 한글 색상 -> 영문 색상명 매핑 (옵션명은 한글, 색상코드표는 영문)
#   긴 단어가 먼저 매칭되도록 길이 내림차순으로 검사한다.
# ---------------------------------------------------------------------------
KR2EN_COLOR = {
    "블랙": "BLACK", "검정": "BLACK", "까망": "BLACK", "흑": "BLACK",
    "화이트": "WHITE", "흰": "WHITE", "백": "WHITE",
    "그레이": "GREY", "회색": "GREY", "그레이핑크": "GREY",
    "네이비": "NAVY", "남색": "NAVY",
    "네이버": "NAVY",  # 오타 흔함
    "브라운": "BROWN", "갈색": "BROWN",
    "베이지": "BEIGE",
    "레드": "RED", "빨강": "RED", "버건디": "WINE",
    "와인": "WINE",
    "블루": "BLUE", "파랑": "BLUE",
    "스카이블루": "SKYBLUE", "스카이": "SKY", "하늘": "SKY",
    "그린": "GREEN", "초록": "GREEN", "녹색": "GREEN",
    "카키": "KHAKI",
    "옐로우": "YELLOW", "노랑": "YELLOW", "옐로": "YELLOW",
    "오렌지": "ORANGE", "주황": "ORANGE",
    "핑크": "PINK", "분홍": "PINK",
    "퍼플": "PURPLE", "보라": "PURPLE",
    "바이올렛": "VIOLET",
    "민트": "MINT",
    "골드": "GOLD", "금": "GOLD",
    "실버": "SILVER", "은": "SILVER",
    "아이보리": "IVORY",
    "크림": "CREAM",
    "코랄": "CORAL", "산호": "CORAL",
    "차콜": "CHARCOAL", "챠콜": "CHARCOAL",
    "올리브": "OLIVE",
    "라임": "LIME",
    "레몬": "LEMON",
    "머스타드": "MUSTARD",
    "스킨": "SKIN", "살구": "APRICOT", "애프리콧": "APRICOT",
    "아쿠아": "AQUA",
    "인디고": "INDIGO",
    "밀리터리": "MILITARY",
    "모스": "MOSS",
    "오트밀": "OATMEAL",
    "틸": "TEAL",
    "커피": "COFFEE",
    "로즈골드": "GOLD",
    "멀티": "MULTIPLE", "혼합": "MULTIPLE",
}
# 길이 내림차순(긴 한글부터)
KR_COLOR_KEYS = sorted(KR2EN_COLOR.keys(), key=len, reverse=True)

# 제조사 정의 -------------------------------------------------------------
MAKERS = {
    "준디자인": {"tag": "J", "label": "준디자인"},
    "자유교역": {"tag": "T", "label": "자유교역"},
}
DEFAULT_BRANDS = ["에버라스트"]

# F~O 고정 기본값 (template H-E 아쿠아 예시 기준) ---------------------------
DEFAULT_FIXED = {
    "F_내용량": 1,
    "G_내용량단위": "086007",
    "H_주요판매국가": "KR",
    "I_가로": 32,
    "J_세로": 23,
    "K_높이": 12,
    "L_총중량": 1000,
    "M_제조국": "CN",
    "N_과세형태": "과세",
    "O_수입여부": 2,
    "Q_KC인증": "",       # 대상아님
    "R": "Y",
    "T_상품분류고정": None,
}

# ---------------------------------------------------------------------------
# 옵션명 파서: "블랙 250", "남성네이비 250 1)슬리퍼형", "그린 230-235" 등
#   반환: dict(gender, color_en, color_kr, size, raw, ok, note)
# ---------------------------------------------------------------------------
def parse_option(raw):
    res = {"raw": raw, "gender": "", "color_kr": "", "color_en": "",
           "size": "", "ok": True, "note": ""}
    if raw is None:
        res["ok"] = False
        res["note"] = "옵션명 없음"
        return res
    s = str(raw).strip()
    size = ""

    # (A) "230(225-230)" 처럼 괄호 앞에 대표 사이즈가 있는 경우 우선 채택
    m_lead = re.match(r"^\s*(\d{3})\s*\(", s)
    if m_lead and 100 <= int(m_lead.group(1)) <= 320:
        size = m_lead.group(1)

    # 메모성 꼬리/괄호 제거
    s_clean = re.split(r"\s*-\s*", s)[0]        # ' - 메모' 제거
    s_clean = re.sub(r"\([^)]*\)", " ", s_clean) # 괄호 메모 제거
    s_clean = re.sub(r"\d\)\S*", " ", s_clean)   # '1)슬리퍼형' 제거
    s_clean = re.sub(r"\s+", " ", s_clean).strip()

    # 성별
    if s_clean.startswith("여성"):
        res["gender"] = "여성"
        s_clean = s_clean[2:]
    elif s_clean.startswith("남성"):
        res["gender"] = "남성"
        s_clean = s_clean[2:]

    # 사이즈 (괄호 앞에서 이미 잡았으면 건너뜀)
    if not size:
        rng = re.search(r"(\d{3})\s*-\s*(\d{3})", s_clean)   # 230-235
        if rng:
            size = rng.group(1)
            res["note"] += f"범위({rng.group(0)})→{size} 사용. "
            s_clean = s_clean.replace(rng.group(0), " ")
        else:
            nums = re.findall(r"\d{2,3}", s_clean)
            cand = [n for n in nums if 100 <= int(n) <= 320]
            if cand:
                size = cand[-1]
                for n in nums:
                    if n == size:
                        s_clean = s_clean.replace(n, " ", 1)
                        break
    # FREE 사이즈 처리 (신발코드표에 FREE 존재)
    if not size and re.search(r"\bfree\b", s, re.IGNORECASE):
        if "FREE" in SHOE_SIZE2CODE:
            size = "FREE"
    if size and size not in SHOE_SIZE2CODE:
        res["note"] += f"사이즈 {size} 코드표에 없음. "
    res["size"] = size

    # 색상: 남은 텍스트에서 한글 색상 토큰 탐색 (긴 것 우선)
    color_kr = ""
    color_en = ""
    txt = s_clean.replace(" ", "")
    for kr in KR_COLOR_KEYS:
        if kr in txt:
            color_kr = kr
            color_en = KR2EN_COLOR[kr]
            break
    res["color_kr"] = color_kr
    res["color_en"] = color_en

    # 판정
    if not size or size not in SHOE_SIZE2CODE:
        res["ok"] = False
        if not size:
            res["note"] += "사이즈 식별 실패. "
    if not color_en or color_en not in COLOR_NAME2CODE:
        res["ok"] = False
        if not color_kr:
            res["note"] += "색상 식별 실패. "
    return res


# ---------------------------------------------------------------------------
# 상세상품명 / 모델명 빌더
#   상품명 예: "X-E 죠깅"  →  알파벳-알파벳 + 한글
#   상세상품명 = 제조사 + 브랜드 + "알파벳-알파벳 (J|T) 한글"
#   모델명     = 상세상품명에서 제조사/브랜드 제외 = "알파벳-알파벳 (J|T) 한글"
# ---------------------------------------------------------------------------
def split_product_name(name):
    """'X-E 죠깅' -> ('X-E', '죠깅')  / 코드부와 한글부 분리.
    'X-E J 보다이'처럼 코드 뒤에 J/T 태그가 이미 있으면 제거한다."""
    s = str(name).strip()
    # 메모 꼬리 제거
    s = re.split(r"\s*-\s*(?=[^A-Za-z])", s, maxsplit=0)[0] if " - " in s else s
    s = s.strip()
    m = re.match(r"^([A-Za-z]+\s*-\s*[A-Za-z0-9]+)\s*(.*)$", s)
    if m:
        code = re.sub(r"\s*-\s*", "-", m.group(1)).upper()
        rest = m.group(2).strip()
        # 한글부 앞에 단독 J/T 태그가 붙어있으면 제거 (중복 방지)
        rest = re.sub(r"^[JT]\s+", "", rest)
        return code, rest
    # 형식이 안 맞으면 통째로 코드부로
    return s, ""


def build_names(prod_name, maker_label, brand, maker_tag):
    code, rest = split_product_name(prod_name)
    # 모델명: 코드 + 태그 + 한글
    parts = [code, maker_tag]
    if rest:
        parts.append(rest)
    model = " ".join(parts)
    detail = f"{maker_label} {brand} {model}".strip()
    return detail, model


# ---------------------------------------------------------------------------
# detailinfo 읽기: 등록상품명(2), 등록옵션명(13). 상품명은 첫행만, 이후 빈칸은 직전 상품 소속
# ---------------------------------------------------------------------------
@st.cache_data
def read_detailinfo(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Template" in wb.sheetnames:
        ws = wb["Template"]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = []
    cur_name = None
    for r in range(5, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        opt = ws.cell(r, 13).value
        if nm:
            cur_name = str(nm).strip()
        if opt is None and nm is None:
            continue
        if opt is None:
            # 상품명만 있고 옵션 없는 행 -> 옵션없는 상품으로 1행
            rows.append({"product": cur_name, "option": ""})
        else:
            rows.append({"product": cur_name, "option": str(opt).strip()})
    return rows


# ---------------------------------------------------------------------------
# 결과 파일 생성
#   원본 템플릿(template_base.xls)을 베이스로 '업로드 템플릿' 시트의
#   데이터 행만 채운 뒤, LibreOffice로 .xls(MS Excel 97)로 변환한다.
#   (GTIN 발급 사이트가 .xls 업로드만 허용하기 때문)
# ---------------------------------------------------------------------------
import subprocess
import tempfile
import shutil

TEMPLATE_BASE = os.path.join(BASE_DIR, "template_base.xls")
DATA_START_ROW = 6   # 업로드 템플릿 시트에서 데이터가 시작되는 행(예시행 위치)


def _find_soffice():
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def _fill_template(df, fixed, base_xlsx_path):
    """베이스 xlsx의 '업로드 템플릿' 시트에 데이터 채우기"""
    wb = openpyxl.load_workbook(base_xlsx_path)
    ws = wb["업로드 템플릿"]

    # 예시행(DATA_START_ROW) 및 그 아래 기존 잔여행 비우기
    for r in range(DATA_START_ROW, ws.max_row + 1):
        for c in range(1, 28):
            ws.cell(r, c).value = None

    for i, row in enumerate(df.itertuples(index=False)):
        r = DATA_START_ROW + i
        vals = [
            i + 1,                    # A 순번
            row.상품분류코드,          # B
            row.제조사명,              # C
            row.브랜드명,              # D
            row.상세상품명,            # E
            fixed["F_내용량"],         # F
            fixed["G_내용량단위"],      # G
            fixed["H_주요판매국가"],    # H
            fixed["I_가로"],           # I
            fixed["J_세로"],           # J
            fixed["K_높이"],           # K
            fixed["L_총중량"],         # L
            fixed["M_제조국"],         # M
            fixed["N_과세형태"],        # N
            fixed["O_수입여부"],        # O
            row.모델명,                # P
            fixed["Q_KC인증"],         # Q
            fixed["R"],                # R
            "",                        # S
            "",                        # T
            row.색상코드,              # U
            row.사이즈코드,            # V
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c).value = v
    return wb


@st.cache_resource
def _base_as_xlsx():
    """원본 template_base.xls를 한 번만 xlsx로 변환해 캐싱"""
    soffice = _find_soffice()
    if soffice is None:
        return None
    tmpdir = tempfile.mkdtemp()
    subprocess.run([soffice, "--headless", "--convert-to", "xlsx",
                    TEMPLATE_BASE, "--outdir", tmpdir],
                   check=True, timeout=120,
                   stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out = os.path.join(tmpdir, "template_base.xlsx")
    return out if os.path.exists(out) else None


def build_output_xls(df, fixed):
    """반환: (xls_bytes 또는 None, xlsx_bytes, 에러메시지)"""
    base_xlsx = _base_as_xlsx()
    if base_xlsx is None:
        return None, None, "LibreOffice를 찾을 수 없어 .xls 변환이 불가합니다."

    wb = _fill_template(df, fixed, base_xlsx)

    tmpdir = tempfile.mkdtemp()
    filled_xlsx = os.path.join(tmpdir, "GTIN_filled.xlsx")
    wb.save(filled_xlsx)
    with open(filled_xlsx, "rb") as f:
        xlsx_bytes = f.read()

    # xlsx -> xls 변환
    soffice = _find_soffice()
    try:
        subprocess.run([soffice, "--headless", "--convert-to", "xls",
                        filled_xlsx, "--outdir", tmpdir],
                       check=True, timeout=120,
                       stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        xls_path = os.path.join(tmpdir, "GTIN_filled.xls")
        if os.path.exists(xls_path):
            with open(xls_path, "rb") as f:
                return f.read(), xlsx_bytes, ""
        return None, xlsx_bytes, ".xls 변환 결과 파일이 생성되지 않았습니다."
    except Exception as e:
        return None, xlsx_bytes, f".xls 변환 오류: {e}"


# ===========================================================================
# UI
# ===========================================================================
st.title("📦 GTIN 코드 발급 양식 생성기")
st.caption("쿠팡 detailinfo → GTIN 업로드 템플릿 변환")

with st.sidebar:
    st.header("1. 공통 설정")

    maker_key = st.radio("제조사명", list(MAKERS.keys()), index=0)
    maker = MAKERS[maker_key]

    # 브랜드(추가 가능)
    if "brands" not in st.session_state:
        st.session_state.brands = list(DEFAULT_BRANDS)
    brand = st.selectbox("브랜드명", st.session_state.brands, index=0)
    with st.expander("브랜드 추가"):
        new_b = st.text_input("새 브랜드명", key="new_brand")
        if st.button("추가") and new_b.strip():
            if new_b.strip() not in st.session_state.brands:
                st.session_state.brands.append(new_b.strip())
                st.rerun()

    st.divider()
    st.header("2. 상품분류 선택 (신발)")
    q = st.text_input("분류 검색(키워드)", placeholder="예: 여성 운동화, 샌들, 남성 부츠")
    if q.strip():
        kws = q.strip().split()
        filtered = [lbl for lbl in CLASS_LABEL2CODE
                    if all(k.lower() in lbl.lower() for k in kws)]
    else:
        filtered = list(CLASS_LABEL2CODE.keys())
    st.caption(f"{len(filtered)}개 매칭")
    if filtered:
        sel_class = st.selectbox("상품분류", filtered)
        class_code = CLASS_LABEL2CODE[sel_class]
    else:
        st.warning("매칭되는 분류가 없습니다.")
        class_code = ""

st.subheader("3. detailinfo 업로드")
up = st.file_uploader("쿠팡 detailinfo 엑셀(.xlsx)", type=["xlsx"])

# 고정값 패널 -----------------------------------------------------------------
with st.expander("⚙️ 고정값(F~O 등) — 기본값은 예시(H-E 아쿠아) 기준, 수정 가능"):
    c1, c2, c3, c4 = st.columns(4)
    fixed = dict(DEFAULT_FIXED)
    fixed["F_내용량"] = c1.number_input("내용량(F)", value=1, step=1)
    fixed["G_내용량단위"] = c2.text_input("내용량단위코드(G)", value="086007")
    fixed["H_주요판매국가"] = c3.text_input("주요판매국가(H)", value="KR")
    fixed["M_제조국"] = c4.text_input("제조국(M)", value="CN")
    c5, c6, c7, c8 = st.columns(4)
    fixed["I_가로"] = c5.number_input("가로cm(I)", value=32, step=1)
    fixed["J_세로"] = c6.number_input("세로cm(J)", value=23, step=1)
    fixed["K_높이"] = c7.number_input("높이cm(K)", value=12, step=1)
    fixed["L_총중량"] = c8.number_input("총중량g(L)", value=1000, step=50)
    c9, c10, c11, c12 = st.columns(4)
    fixed["N_과세형태"] = c9.text_input("과세형태(N)", value="과세")
    fixed["O_수입여부"] = c10.number_input("수입여부(O)", value=2, step=1)
    fixed["Q_KC인증"] = c11.text_input("KC인증(Q)", value="")
    fixed["R"] = c12.text_input("R열", value="Y")

if up is not None and class_code:
    rows = read_detailinfo(up.getvalue())
    st.success(f"옵션 {len(rows):,}행 로드 / 고유 상품 "
               f"{len(set(r['product'] for r in rows)):,}개")

    # 상품 단위 선택 ---------------------------------------------------------
    products = sorted(set(r["product"] for r in rows if r["product"]))
    st.subheader("4. 변환할 상품 선택")
    sel_products = st.multiselect("상품(미선택 시 전체)", products)
    target = sel_products if sel_products else products

    # 각 옵션 파싱 -----------------------------------------------------------
    parsed_rows = []
    for r in rows:
        if r["product"] not in target:
            continue
        p = parse_option(r["option"])
        detail, model = build_names(r["product"], maker["label"], brand, maker["tag"])
        color_code = COLOR_NAME2CODE.get(p["color_en"], "")
        size_code = SHOE_SIZE2CODE.get(p["size"], "")
        parsed_rows.append({
            "상품명": r["product"],
            "옵션명": r["option"],
            "상품분류코드": class_code,
            "제조사명": maker["label"],
            "브랜드명": brand,
            "상세상품명": detail,
            "모델명": model,
            "색상(인식)": p["color_kr"],
            "색상영문": p["color_en"],
            "색상코드": color_code,
            "사이즈": p["size"],
            "사이즈코드": size_code,
            "성별": p["gender"],
            "상태": "OK" if (color_code and size_code) else "확인필요",
            "메모": p["note"],
        })

    df = pd.DataFrame(parsed_rows)
    if df.empty:
        st.info("선택된 상품의 옵션이 없습니다.")
        st.stop()

    ok_n = (df["상태"] == "OK").sum()
    ng_n = (df["상태"] == "확인필요").sum()
    st.subheader("5. 매핑 결과 검토 / 수동 보정")
    st.write(f"✅ 자동매핑 성공 **{ok_n}** / ⚠️ 확인필요 **{ng_n}**")

    # 색상/사이즈 선택지(라벨)
    color_options = [""] + [f'{c["code"]} | {c["name"]}' for c in COLOR_LIST]
    size_options = [""] + [f'{s["code"]} | {s["size"]}' for s in SHOE_LIST]
    color_code2label = {c["code"]: f'{c["code"]} | {c["name"]}' for c in COLOR_LIST}
    size_code2label = {s["code"]: f'{s["code"]} | {s["size"]}' for s in SHOE_LIST}

    # 편집용 컬럼 추가
    df_edit = df.copy()
    df_edit["색상코드_선택"] = df_edit["색상코드"].map(
        lambda x: color_code2label.get(x, ""))
    df_edit["사이즈코드_선택"] = df_edit["사이즈코드"].map(
        lambda x: size_code2label.get(x, ""))

    edited = st.data_editor(
        df_edit[["상품명", "옵션명", "상태", "메모",
                 "상세상품명", "모델명",
                 "색상(인식)", "색상코드_선택",
                 "사이즈", "사이즈코드_선택"]],
        column_config={
            "색상코드_선택": st.column_config.SelectboxColumn(
                "색상코드", options=color_options, width="medium"),
            "사이즈코드_선택": st.column_config.SelectboxColumn(
                "사이즈코드", options=size_options, width="small"),
            "상품명": st.column_config.TextColumn(disabled=True),
            "옵션명": st.column_config.TextColumn(disabled=True),
            "상태": st.column_config.TextColumn(disabled=True),
            "메모": st.column_config.TextColumn(disabled=True),
            "상세상품명": st.column_config.TextColumn(disabled=True),
            "모델명": st.column_config.TextColumn(disabled=True),
            "색상(인식)": st.column_config.TextColumn(disabled=True),
            "사이즈": st.column_config.TextColumn(disabled=True),
        },
        hide_index=True,
        use_container_width=True,
        height=500,
        key="editor",
    )

    # 선택값 -> 코드 반영
    def label2code(label):
        if not label:
            return ""
        return str(label).split(" | ")[0].strip()

    final = df.copy()
    final["색상코드"] = edited["색상코드_선택"].map(label2code).values
    final["사이즈코드"] = edited["사이즈코드_선택"].map(label2code).values

    remaining = ((final["색상코드"] == "") | (final["사이즈코드"] == "")).sum()
    if remaining:
        st.warning(f"아직 코드가 비어있는 행 {remaining}개 — 위 표에서 직접 선택하세요.")

    st.subheader("6. 다운로드")
    with st.spinner("GTIN 템플릿(.xls) 생성 중…"):
        xls_bytes, xlsx_bytes, err = build_output_xls(final, fixed)

    if xls_bytes:
        st.success("원본 템플릿 기반 .xls 파일이 생성됐습니다.")
        st.download_button(
            "📥 GTIN 템플릿(.xls) 다운로드  ← 업로드용",
            data=xls_bytes,
            file_name="GTIN_upload.xls",
            mime="application/vnd.ms-excel",
            type="primary",
        )
    else:
        st.error(f".xls 생성 실패: {err}\n\n"
                 "아래 .xlsx로 받아서 엑셀에서 '다른 이름으로 저장 → "
                 "Excel 97-2003 통합 문서(*.xls)'로 변환해 업로드하세요.")
    if xlsx_bytes:
        st.download_button(
            "📄 (백업) .xlsx 다운로드",
            data=xlsx_bytes,
            file_name="GTIN_upload.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with st.expander("최종 데이터 미리보기"):
        st.dataframe(
            final[["상품명", "옵션명", "상품분류코드", "제조사명", "브랜드명",
                   "상세상품명", "모델명", "색상코드", "사이즈코드"]],
            use_container_width=True, hide_index=True)
else:
    st.info("좌측에서 제조사·브랜드·상품분류를 설정하고 detailinfo 파일을 업로드하세요.")
